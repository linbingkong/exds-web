from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
import pandas as pd
import io
from datetime import datetime
from urllib.parse import quote
from webapp.models.contract import Contract, ContractCreate, ContractListResponse, calculate_contract_status
from webapp.services.contract_service import ContractService
from webapp.tools.mongo import DATABASE
from webapp.tools.security import get_current_active_user, User
from webapp.tools.excel_handler import ExcelReader, DataValidator, ContractDataTransformer
from webapp.api.dependencies.authz import require_permission, CurrentUserContext
from webapp.api.masking import mask_response_for_user
from webapp.services.customer_name_masking_service import customer_name_masking_service

router = APIRouter(prefix="/retail-contracts", tags=["Retail Contracts"])


@router.post("", response_model=dict, status_code=status.HTTP_201_CREATED)
async def create_contract(
    contract: ContractCreate,
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """创建新合同"""
    service = ContractService(DATABASE)
    try:
        result = service.create(
            contract_data=contract.model_dump(exclude_unset=True),
            operator=current_user.username
        )
        return result
    except ValueError as e:
        error_msg = str(e)
        if "不存在" in error_msg or "无效" in error_msg or "状态不是" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )


@router.get("", response_model=ContractListResponse)
async def list_contracts(
    contract_name: Optional[str] = Query(None, description="合同名称（模糊搜索）"),
    package_name: Optional[str] = Query(None, description="套餐名称（模糊搜索）"),
    customer_name: Optional[str] = Query(None, description="客户名称（模糊搜索）"),
    status: Optional[str] = Query(None, description="合同状态（pending/active/expired）"),
    purchase_start_month: Optional[str] = Query(None, description="购电开始月份筛选（yyyy-MM）"),
    purchase_end_month: Optional[str] = Query(None, description="购电结束月份筛选（yyyy-MM）"),
    year: Optional[int] = Query(None, description="年份筛选（覆盖月份范围）"),
    sort_field: Optional[str] = Query("created_at", description="排序字段"),
    sort_order: Optional[str] = Query("desc", description="排序方向(asc/desc)"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页大小"),
    current_user: User = Depends(get_current_active_user),
    ctx: CurrentUserContext = Depends(require_permission("module:customer_retail_contracts:view")),
):
    """
    获取合同列表

    支持筛选：
    - contract_name: 合同名称（模糊搜索）
    - package_name: 套餐名称（模糊搜索）
    - customer_name: 客户名称（模糊搜索）
    - status: 合同状态（pending/active/expired）

    支持分页：
    - page: 页码（从1开始）
    - page_size: 每页数量
    """
    service = ContractService(DATABASE)
    use_masked_customer_search = bool(customer_name and not ctx.can_view_real_customer_name)
    matched_customer_ids = customer_name_masking_service.search_customer_ids_by_keyword(customer_name or "") if use_masked_customer_search else []
    result = service.list(
        filters={
            "contract_name": contract_name,
            "package_name": package_name,
            "customer_name": None if use_masked_customer_search else customer_name,
            "customer_ids": matched_customer_ids,
            "status": status,
            "year": year,
            "purchase_start_month": purchase_start_month,
            "purchase_end_month": purchase_end_month
        },
        page=page,
        page_size=page_size,
        sort_field=sort_field,
        sort_order=sort_order
    )
    return mask_response_for_user(result, ctx)


@router.get("/years", response_model=List[int])
async def get_contract_years(
    current_user: User = Depends(get_current_active_user)
):
    """获取所有合同中涉及的年份"""
    service = ContractService(DATABASE)
    return service.get_available_years()


@router.post("/import", summary="导入合同数据")
async def import_contracts(
    file: UploadFile = File(..., description="交易中心平台下载的Excel文件"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """
    批量导入合同数据（同步方案）

    Excel文件格式要求：
    - 必需列：套餐, 购买用户, 购买电量, 购买时间-开始, 购买时间-结束
    - 忽略列：序号, 代理销售费模型, 签章状态
    - 套餐名称必须存在于系统中且状态为已生效
    - 客户名称必须存在于系统中且状态为正常
    - 日期格式：YYYY-MM 或 YYYY-MM-DD
    - 购买电量必须大于0
    - 购买结束月份必须 >= 购买开始月份

    返回：
    - total: 总行数
    - success: 成功导入的数量
    - failed: 失败的数量
    - errors: 错误详情列表
    """
    try:
        # 1. 读取Excel文件
        contents = await file.read()

        # 2. 初始化处理组件
        excel_reader = ExcelReader()
        validator = DataValidator(DATABASE)
        transformer = ContractDataTransformer(DATABASE)

        # 3. 读取并验证文件结构
        df = excel_reader.read_excel_file(contents)
        excel_reader.validate_excel_structure(df)

        # 4. 逐行校验和导入
        success_count = 0
        failed_count = 0
        errors = []

        for index, row in df.iterrows():
            row_number = index + 2  # Excel从1开始，且有表头

            try:
                # 解析行数据
                row_data = excel_reader.parse_row_data(row, row_number)

                # 执行多层校验
                validation_errors = []
                validation_errors.extend(validator.validate_required_fields(row_data))
                validation_errors.extend(validator.validate_related_data(row_data))
                validation_errors.extend(validator.validate_business_rules(row_data))

                if validation_errors:
                    errors.extend(validation_errors)
                    failed_count += 1
                    continue

                # 转换数据格式
                contract_data = transformer.transform_row_to_contract(row_data, current_user.username)

                # 检查合同唯一性
                uniqueness_errors = validator.validate_contract_uniqueness(
                    contract_data['customer_id'],
                    contract_data['purchase_start_month'],
                    contract_data['purchase_end_month'],
                    row_number
                )

                if uniqueness_errors:
                    errors.extend(uniqueness_errors)
                    failed_count += 1
                    continue

                # 插入数据库
                DATABASE.retail_contracts.insert_one(contract_data)
                success_count += 1

            except Exception as e:
                errors.append({
                    'row': row_number,
                    'field': 'general',
                    'value': None,
                    'message': str(e),
                    'suggestion': '请检查该行数据的完整性和格式'
                })
                failed_count += 1

        # 5. 返回结果
        return {
            "total": len(df),
            "success": success_count,
            "failed": failed_count,
            "errors": errors
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"导入失败：{str(e)}")


@router.get("/export", summary="导出合同数据")
async def export_contracts(
    package_name: Optional[str] = Query(None, description="套餐名称筛选"),
    customer_name: Optional[str] = Query(None, description="客户名称筛选"),
    status: Optional[str] = Query(None, description="合同状态筛选"),
    start_month: Optional[str] = Query(None, description="购电开始月份筛选(YYYY-MM)"),
    end_month: Optional[str] = Query(None, description="购电结束月份筛选(YYYY-MM)"),
    current_user: User = Depends(get_current_active_user),
    ctx: CurrentUserContext = Depends(require_permission("customer:contract:export"))
):
    """
    导出合同数据为Excel文件

    支持筛选：
    - package_name: 套餐名称（模糊搜索）
    - customer_name: 客户名称（模糊搜索）
    - status: 合同状态（pending/active/expired）
    - start_month: 购电开始月份筛选
    - end_month: 购电结束月份筛选
    """
    try:
        # 1. 构建查询条件
        query = {}
        if package_name:
            query["package_name"] = {"$regex": package_name, "$options": "i"}
        if customer_name and ctx.can_view_real_customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        elif customer_name:
            matched_customer_ids = customer_name_masking_service.search_customer_ids_by_keyword(customer_name)
            if matched_customer_ids:
                query["customer_id"] = {"$in": matched_customer_ids}
            else:
                query["customer_id"] = {"$in": ["__no_match__"]}
        if start_month:
            query["purchase_start_month"] = {"$gte": start_month}
        if end_month:
            query["purchase_end_month"] = {"$lte": end_month}

        # 2. 查询数据
        contracts = list(DATABASE.retail_contracts.find(query).sort("created_at", -1))
        contracts = mask_response_for_user(contracts, ctx)

        # 3. 计算虚拟状态并添加到数据中
        processed_contracts = []
        for contract in contracts:
            status_value = calculate_contract_status(
                contract.get("purchase_start_month"),
                contract.get("purchase_end_month")
            )

            # 应用状态筛选
            if status and status != "all":
                if status_value != status:
                    continue

            # 格式化数据
            formatted_contract = {
                '合同编号': str(contract.get('_id', '')),
                '合同名称': contract.get('contract_name', ''),
                '套餐名称': contract.get('package_name', ''),
                '购买用户': contract.get('customer_name', ''),
                '购买电量': contract.get('purchasing_electricity_quantity', 0),
                '购电开始月份': _format_date(contract.get('purchase_start_month'), 'date'),
                '购电结束月份': _format_date(contract.get('purchase_end_month'), 'date'),
                '合同状态': _format_status(status_value),
                '创建时间': _format_date(contract.get('created_at'), 'datetime'),
                '更新时间': _format_date(contract.get('updated_at'), 'datetime'),
            }

            processed_contracts.append(formatted_contract)

        # 4. 生成Excel文件
        excel_data = _generate_excel_file(processed_contracts, {
            'package_name': package_name,
            'customer_name': customer_name,
            'status': status,
            'start_month': start_month,
            'end_month': end_month
        })

        # 5. 返回文件流
        filename = f"零售合同数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # 对文件名进行 URL 编码以支持中文
        encoded_filename = quote(filename)

        return StreamingResponse(
            io.BytesIO(excel_data.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败：{str(e)}")


@router.get("/{contract_id}", response_model=dict)
async def get_contract(
    contract_id: str,
    current_user: User = Depends(get_current_active_user),
    ctx: CurrentUserContext = Depends(require_permission("module:customer_retail_contracts:view")),
):
    """获取合同详情"""
    service = ContractService(DATABASE)
    try:
        result = service.get_by_id(contract_id)
        return mask_response_for_user(result, ctx)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


@router.put("/{contract_id}", response_model=dict)
async def update_contract(
    contract_id: str,
    contract: ContractCreate,
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """更新合同"""
    service = ContractService(DATABASE)
    try:
        result = service.update(
            contract_id=contract_id,
            contract_data=contract.model_dump(exclude_unset=True),
            operator=current_user.username
        )
        return result
    except ValueError as e:
        error_msg = str(e)
        if "不存在" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=error_msg
            )
        elif "状态" in error_msg or "无效" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )


@router.delete("/{contract_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_contract(
    contract_id: str,
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("customer:contract:delete"))
):
    """删除合同（仅待生效状态）"""
    service = ContractService(DATABASE)
    try:
        service.delete(contract_id)
        return None  # 204 No Content
    except ValueError as e:
        error_msg = str(e)
        if "状态" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=error_msg
            )


def _format_date(date_value, format_type='date'):
    """格式化日期"""
    if date_value is None:
        return ""

    if isinstance(date_value, str):
        date_value = datetime.fromisoformat(date_value.replace('Z', '+00:00'))

    if format_type == 'date':
        return date_value.strftime("%Y-%m")
    elif format_type == 'datetime':
        return date_value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return str(date_value)


def _format_status(status):
    """格式化合同状态为中文"""
    status_map = {
        'pending': '待生效',
        'active': '生效',
        'expired': '已过期'
    }
    return status_map.get(status, status)


def _generate_excel_file(data, params):
    """生成Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 创建DataFrame
    if not data:
        # 创建空表格带列名
        columns = [
            '合同编号', '合同名称', '套餐名称', '购买用户', '购买电量',
            '购电开始月份', '购电结束月份', '合同状态', '创建时间', '更新时间'
        ]
        df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(data)

    # 创建Excel工作簿
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='合同数据')

        # 获取工作表
        worksheet = writer.sheets['合同数据']

        # 设置样式
        _format_excel_worksheet(worksheet, df, params)

    output.seek(0)
    return output


def _format_excel_worksheet(worksheet, df, params):
    """格式化Excel工作表"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 设置标题行样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')

    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用标题行样式
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 应用数据行样式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # 状态列颜色标记
            if cell.column_letter == 'H':  # 状态列
                if cell.value == '生效':
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                elif cell.value == '已过期':
                    cell.fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
                elif cell.value == '待生效':
                    cell.fill = PatternFill(start_color='E8F4FF', end_color='E8F4FF', fill_type='solid')

    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    worksheet.freeze_panes = 'A2'

    # 添加筛选器
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    # 添加页眉信息
    if not df.empty:
        # 在顶部添加筛选条件说明
        worksheet.insert_rows(1)
        worksheet.cell(row=1, column=1, value='导出条件：')

        filter_desc = _build_filter_description(params)
        worksheet.cell(row=1, column=2, value=filter_desc)

        # 合并单元格
        worksheet.merge_cells(f'B1:{get_column_letter(worksheet.max_column)}1')

        # 设置页眉样式
        header_cell = worksheet.cell(row=1, column=1)
        header_cell.font = Font(name='微软雅黑', size=10, bold=True)
        header_cell.alignment = Alignment(horizontal='left', vertical='center')


def _build_filter_description(params):
    """构建筛选条件描述"""
    descriptions = []

    if params.get('package_name'):
        descriptions.append(f"套餐名称：{params['package_name']}")

    if params.get('customer_name'):
        descriptions.append(f"客户名称：{params['customer_name']}")

    if params.get('status') and params['status'] != 'all':
        status_map = {'pending': '待生效', 'active': '生效', 'expired': '已过期'}
        descriptions.append(f"合同状态：{status_map.get(params['status'], params['status'])}")

    if params.get('start_month') or params.get('end_month'):
        start = params.get('start_month', '开始')
        end = params.get('end_month', '结束')
        descriptions.append(f"购电时间：{start} ~ {end}")

    return '；'.join(descriptions) if descriptions else '无筛选条件'


# ##############################################################################
# 合同PDF文件管理API (Contract PDF APIs)
# ##############################################################################

from webapp.services.contract_pdf_service import ContractPdfService


@router.post("/upload-pdfs", summary="批量上传合同PDF文件")
async def upload_contract_pdfs(
    files: List[UploadFile] = File(..., description="PDF文件列表"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """
    批量上传PDF文件，自动匹配合同记录
    
    文件命名规范：客户名称-合同描述.pdf
    例如：富联精密科技（赣州）有限公司-26年零售平台电子合同.pdf
    
    匹配规则：
    1. 根据客户名称模糊匹配
    2. 如果描述中有年份(如"26年"表示2026年)，用年份进一步过滤
    3. 匹配单个且无已上传PDF -> 自动导入
    4. 匹配多个 -> 需要用户确认
    5. 匹配单个但已有PDF -> 需要用户确认是否覆盖
    
    返回:
    - matched: 自动匹配并保存的文件列表
    - pending: 需要用户确认的文件列表
    - errors: 处理错误列表
    """
    pdf_service = ContractPdfService(DATABASE)
    
    matched = []
    pending = []  # 需要用户确认的
    errors = []
    
    for file in files:
        # 验证文件类型
        if not file.filename.lower().endswith('.pdf'):
            errors.append({
                "filename": file.filename,
                "error": "不是PDF文件"
            })
            continue
        
        try:
            # 读取文件内容
            content = await file.read()
            
            # 匹配合同记录
            match_result = pdf_service.match_pdf_to_contracts(file.filename)
            
            if match_result["auto_import"] and match_result["target_contract"]:
                # 可以自动导入
                contract = match_result["target_contract"]
                success = pdf_service.save_pdf_to_contract(
                    contract_id=contract["_id"],
                    pdf_data=content,
                    filename=file.filename,
                    uploader=current_user.username
                )
                
                if success:
                    matched.append({
                        "filename": file.filename,
                        "contract_id": contract["_id"],
                        "contract_name": contract["contract_name"],
                        "customer_name": contract["customer_name"]
                    })
                else:
                    errors.append({
                        "filename": file.filename,
                        "error": "保存失败"
                    })
            else:
                # 需要用户确认
                pending.append({
                    "filename": file.filename,
                    "reason": match_result["reason"],
                    "candidates": match_result["matches"],
                    "target_contract": match_result["target_contract"]
                })
                
        except Exception as e:
            errors.append({
                "filename": file.filename,
                "error": str(e)
            })
    
    return {
        "matched": matched,
        "pending": pending,  # 改名为pending更准确
        "errors": errors,
        "summary": {
            "total": len(files),
            "matched_count": len(matched),
            "pending_count": len(pending),
            "error_count": len(errors)
        }
    }


from webapp.tools.contract_analyzer import ContractAnalyzer
from webapp.models.pdf_import import ParsePdfResponse, ImportCreateRequest
from webapp.services.customer_service import CustomerService
from webapp.services.package_service import PackageService
import dateutil.parser


@router.post("/parse-pdf", response_model=ParsePdfResponse, summary="解析合同PDF进行预览")
async def parse_contract_pdf(
    file: UploadFile = File(..., description="合同原件PDF文件"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """
    解析PDF返回其中提取的数据，并检查与现有系统的重复情况。
    """
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="请上传PDF格式的文件"
        )
    
    analyzer = ContractAnalyzer()
    content = await file.read()

    try:
        parsed_data = analyzer.analyze(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    customer_name = parsed_data.get("customer_name")
    if not customer_name:
        raise HTTPException(status_code=400, detail="未能从PDF中解析出客户名称")
    
    # 地区推断(来自简称生成方法的地名推断逻辑也可复用，此处ContractAnalyzer已经根据地名生成简称，我们可以取其首部地名作为location)
    location = None
    regions = [
        '景德镇', '南昌', '九江', '赣州', '吉安', '宜春', '抚州', '上饶', 
        '萍乡', '新余', '鹰潭', '上高', '丰城', '峡江', '新干', '高安', 
        '井冈山', '宜丰', '青云谱', '江西'
    ]
    for r in sorted(regions, key=len, reverse=True):
        if r in customer_name:
            location = r
            break
            
    # 检查数据库
    customer_service = CustomerService(DATABASE)
    existing_customer = customer_service.collection.find_one({"user_name": customer_name})
    is_customer_new = existing_customer is None

    package_name = parsed_data.get("package_name")
    is_package_new = False
    if package_name:
        # Check if package exists
        package_doc = DATABASE.retail_packages.find_one({"package_name": package_name})
        if not package_doc:
            is_package_new = True
            
    # 检查合同是否重复（同客户名，同起始时间）
    is_contract_duplicate = False
    duplicate_contract_id = None
    period_str = parsed_data.get("period")
    if period_str and existing_customer:
        # 尝试简易提取开始时间，例如 "2024年1月至2024年12月" -> "2024年1月"
        start_part = period_str.split("至")[0].strip()
        try:
            # 转换为日期以进行查询
            import re
            m = re.search(r'(\d{4})年(\d{1,2})月', start_part)
            if m:
                year, month = int(m.group(1)), int(m.group(2))
                start_date = datetime(year, month, 1)
                
                # 查询该客户在此月是否已有合同
                dup = DATABASE.retail_contracts.find_one({
                    "customer_name": customer_name,
                    "purchase_start_month": {"$gte": start_date, "$lt": datetime(year, month+1, 1) if month < 12 else datetime(year+1, 1, 1)}
                })
                if dup:
                    is_contract_duplicate = True
                    duplicate_contract_id = str(dup["_id"])
        except:
            pass

    return ParsePdfResponse(
        customer_name=customer_name,
        customer_short_name=parsed_data.get("customer_short_name"),
        period=period_str,
        package_name=package_name,
        total_electricity=parsed_data.get("total_electricity"),
        attachment2=parsed_data.get("attachment2", []),
        location=location,
        is_customer_new=is_customer_new,
        is_package_new=is_package_new,
        is_contract_duplicate=is_contract_duplicate,
        duplicate_contract_id=duplicate_contract_id
    )


@router.post("/import-create", summary="确认导入创建合同及相关数据")
async def import_create_contract(
    req: ImportCreateRequest,
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """
    根据前端确认的数据创建客户、套餐和合同
    """
    # 1. 处理客户
    customer_service = CustomerService(DATABASE)
    existing_customer = customer_service.collection.find_one({"user_name": req.customer_name})
    
    # 构建新户号数据
    new_accounts = []
    # 预处理附件作为户号
    if req.attachment2:
        # 以account_id进行分组
        account_map = {}
        for mp in req.attachment2:
            acc_id = mp.measuring_point # 从ContractAnalyzer, 我们提取的是户号(meter_id位置)还是计量点
            # 修改: contract_analyzer目前提取的是 `(\d{10,})` as meter_id, `(\d{8,10})` as measuring_point. (通常10位长的是用电户号, 8-10位的是资产号)
            # 不过我们目前仅原样存储
            # 这里按照Account模型组装
            # Account requires account_id, meters, metering_points
            acc_id = mp.meter_id  # 假设10位数字为account id
            if acc_id not in account_map:
                account_map[acc_id] = {
                    "account_id": acc_id,
                    "meters": [],
                    "metering_points": []
                }
            account_map[acc_id]["metering_points"].append({
                "mp_no": mp.measuring_point,
                "mp_name": ""
            })
            
        new_accounts = list(account_map.values())
        
    customer_id = None
    if not existing_customer:
        customer_data = {
            "user_name": req.customer_name,
            "short_name": req.customer_short_name or req.customer_name[:4],
            "location": req.location,
            "accounts": new_accounts
        }
        res = customer_service.create(customer_data, current_user.username)
        customer_id = res["id"]
    else:
        customer_id = str(existing_customer["_id"])
        # 更新location和追加accounts
        updates = {}
        if req.location and not existing_customer.get("location"):
            updates["location"] = req.location
            
        if new_accounts:
            existing_accounts = existing_customer.get("accounts", [])
            existing_acc_ids = {acc["account_id"] for acc in existing_accounts}
            for na in new_accounts:
                if na["account_id"] not in existing_acc_ids:
                    existing_accounts.append(na)
            updates["accounts"] = existing_accounts
            
        if updates:
            customer_service.update(customer_id, updates, current_user.username)
            
    # 2. 处理套餐
    package_service = PackageService(DATABASE)
    package_doc = DATABASE.retail_packages.find_one({"package_name": req.package_name})
    package_id = None
    if not package_doc:
        # 创建默认草稿套餐
        pkg_res = package_service.create({
            "package_name": req.package_name,
            "package_type": "time_based", # 默认分时套餐
            "template_id": None,
            "description": "从PDF导入自动创建的草稿套餐",
            "parameters": {}
        }, current_user.username)
        package_id = pkg_res["id"]
        # 服务内状态默认可能为draft, 如果不是, 可以手动update... 不过RetailPackageCreate目前不强制status
    else:
        package_id = str(package_doc["_id"])
        
    # 3. 处理合同
    import re
    # 解析购电时间 "2024年1月至2024年12月"
    start_date, end_date = None, None
    if req.period:
        m = re.findall(r'(\d{4})年(\d{1,2})月', req.period)
        if len(m) >= 2:
            start_date = datetime(int(m[0][0]), int(m[0][1]), 1)
            end_date = datetime(int(m[1][0]), int(m[1][1]), 1)
        elif len(m) == 1:
            start_date = datetime(int(m[0][0]), int(m[0][1]), 1)
            end_date = start_date
            
    if not start_date or not end_date:
        raise HTTPException(status_code=400, detail="购电起止月份解析失败，请检查PDF时期格式是否标准")
        
    contract_service = ContractService(DATABASE)
    contract_data = {
        "customer_id": customer_id,
        "customer_name": req.customer_name,
        "package_id": package_id,
        "package_name": req.package_name,
        "purchasing_electricity_quantity": req.total_electricity or 0.0,
        "purchase_start_month": start_date,
        "purchase_end_month": end_date,
        # 合同名称自动在 service.create 中基于 customer short_name 和时间生成
    }
    
    contract_res = contract_service.create(contract_data, current_user.username)
    return {
        "success": True,
        "contract_id": contract_res["id"],
        "customer_id": customer_id,
        "package_id": package_id
    }


@router.get("/{contract_id}/pdf", summary="获取合同PDF文件")
async def get_contract_pdf(
    contract_id: str,
    current_user: User = Depends(get_current_active_user),
    ctx: CurrentUserContext = Depends(require_permission("module:customer_retail_contracts:view")),
):
    """
    获取合同的PDF文件，返回PDF文件流供预览/下载
    """
    if not ctx.can_view_real_customer_name:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="当前角色不允许查看原始合同文件",
        )

    pdf_service = ContractPdfService(DATABASE)
    
    result = pdf_service.get_contract_pdf(contract_id)
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该合同未上传PDF文件"
        )
    
    # 返回PDF文件流
    from urllib.parse import quote
    encoded_filename = quote(result["pdf_filename"])
    
    return StreamingResponse(
        io.BytesIO(result["pdf_data"]),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/{contract_id}/upload-pdf", summary="为指定合同上传PDF")
async def upload_single_pdf(
    contract_id: str,
    file: UploadFile = File(..., description="PDF文件"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:customer_retail_contracts:edit"))
):
    """
    为指定合同上传/替换PDF文件
    """
    # 验证文件类型
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="请上传PDF格式的文件"
        )
    
    pdf_service = ContractPdfService(DATABASE)
    
    try:
        content = await file.read()
        
        success = pdf_service.save_pdf_to_contract(
            contract_id=contract_id,
            pdf_data=content,
            filename=file.filename,
            uploader=current_user.username
        )
        
        if success:
            return {
                "success": True,
                "message": "PDF上传成功",
                "contract_id": contract_id,
                "filename": file.filename
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="合同不存在或保存失败"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"上传失败：{str(e)}"
        )


@router.get("/{contract_id}/has-pdf", summary="检查合同是否有PDF")
async def check_contract_has_pdf(
    contract_id: str,
    current_user: User = Depends(get_current_active_user),
    _ctx: CurrentUserContext = Depends(require_permission("module:customer_retail_contracts:view")),
):
    """
    检查合同是否已上传PDF文件
    """
    pdf_service = ContractPdfService(DATABASE)
    has_pdf = pdf_service.has_pdf(contract_id)
    
    return {"has_pdf": has_pdf}
