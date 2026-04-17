import io
from typing import Any, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExportService:
    def __init__(self):
        pass

    def export_wholesale_to_excel(
        self,
        date_str: str,
        version_name: str,
        data: dict,
        export_context: Optional[Dict[str, Any]] = None,
    ) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "批发侧结算详情"

        # 样式定义
        header_font = Font(name='微软雅黑', bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 定义颜色
        colors = {
            'blue': 'E3F2FD',    # 中长期
            'orange': 'FFF3E0',  # 日前
            'green': 'E8F5E9',   # 实时
            'red': 'FFEBEE',     # 电能量
            'purple': 'F3E5F5',  # 标准值
            'year': 'E8EAF6',    # 年度合约
            'month': 'E0F7FA',   # 月度合约
            'intra': 'FFF8E1',   # 月内合约
        }

        export_context = export_context or {}
        market_contract_avg = export_context.get("market_contract_avg", {})
        contract_breakdown = export_context.get("contract_breakdown", {})
        yearly_contract = contract_breakdown.get("年度", {})
        monthly_contract = contract_breakdown.get("月度", {})
        intra_month_contract = contract_breakdown.get("月内", {})

        # 1. 概览信息
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"批发侧结算详情 - {date_str} ({version_name})"
        ws['A1'].font = Font(name='微软雅黑', bold=True, size=14)
        ws['A1'].alignment = center_align

        # 2. 第一级表头
        curr_col = 1
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 35

        # 辅助函数：应用样式
        def apply_style(cell_range, bg_color=None):
            res = ws[cell_range]
            # 如果是单单元格，包装成嵌套元组以统一迭代逻辑
            cells_to_style = ((res,),) if not isinstance(res, (tuple, list)) else res
            
            for row in cells_to_style:
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
                    if bg_color:
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 时段
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.cell(row=2, column=1, value="时段")
        apply_style('A2:A3')
        curr_col = 2

        groups = [
            ("中长期合约电费", 3, colors['blue']),
            ("日前市场偏差", 3, colors['orange']),
            ("实时市场偏差", 3, colors['green']),
            ("电能量", 2, colors['red']),
            ("标准值", 5, colors['purple']),
            ("年度合约", 3, colors['year']),
            ("月度合约", 3, colors['month']),
            ("月内合约", 3, colors['intra']),
        ]

        for text, span, color in groups:
            ws.merge_cells(start_row=2, start_column=curr_col, end_row=2, end_column=curr_col + span - 1)
            cell = ws.cell(row=2, column=curr_col, value=text)
            cell.font = header_font
            range_str = f"{ws.cell(row=2, column=curr_col).coordinate}:{ws.cell(row=2, column=curr_col + span - 1).coordinate}"
            apply_style(range_str, color)
            curr_col += span

        # 3. 第二级表头 (子列名称 + 公式)
        sub_headers = [
            ("合同电量\n①", colors['blue']),
            ("合同均价\n②", colors['blue']),
            ("差价电费\n③=①×(②-⑧)", colors['blue']),
            ("出清电量\n④", colors['orange']),
            ("市场均价\n⑤", colors['orange']),
            ("差价电费\n⑥=④×(⑤-⑧)", colors['orange']),
            ("实际用量\n⑦", colors['green']),
            ("市场均价\n⑧", colors['green']),
            ("全量电费\n⑨=⑦×⑧", colors['green']),
            ("电费合计\n⑩=③+⑥+⑨", colors['red']),
            ("结算均价\n⑪=⑩÷⑦", colors['red']),
            ("机制电量\n⑫", colors['purple']),
            ("签约比例\n⑬=(①+⑫)÷⑦", colors['purple']),
            ("电费合计\n⑭=分段公式", colors['purple']),
            ("结算均价\n⑮=⑭÷⑦", colors['purple']),
            ("全市场合同均价\n⑯", colors['purple']),
            ("合同电量", colors['year']),
            ("合同均价", colors['year']),
            ("差价电费\n=合同电量×(合同均价-⑧)", colors['year']),
            ("合同电量", colors['month']),
            ("合同均价", colors['month']),
            ("差价电费\n=合同电量×(合同均价-⑧)", colors['month']),
            ("合同电量", colors['intra']),
            ("合同均价", colors['intra']),
            ("差价电费\n=合同电量×(合同均价-⑧)", colors['intra']),
        ]

        for i, (text, color) in enumerate(sub_headers):
            cell = ws.cell(row=3, column=i + 2, value=text)
            cell.font = Font(name='微软雅黑', size=9)
            apply_style(cell.coordinate, color)

        # 4. 填充数据
        details = data.get('wholesale_period_details', [])
        for idx, p in enumerate(details):
            row_idx = idx + 4
            year_volumes = yearly_contract.get("volumes", [])
            year_prices = yearly_contract.get("prices", [])
            month_volumes = monthly_contract.get("volumes", [])
            month_prices = monthly_contract.get("prices", [])
            intra_volumes = intra_month_contract.get("volumes", [])
            intra_prices = intra_month_contract.get("prices", [])
            market_prices = market_contract_avg.get("prices", [])

            # A: 时段
            ws.cell(row=row_idx, column=1, value=p.get('period'))
            
            # B/C/D: 中长期合约由年度、月度、月内三类合同汇总计算
            ws.cell(row=row_idx, column=2, value=f"=R{row_idx}+U{row_idx}+X{row_idx}")
            ws.cell(
                row=row_idx,
                column=3,
                value=(
                    f"=IF(B{row_idx}=0,0,"
                    f"(R{row_idx}*S{row_idx}+U{row_idx}*V{row_idx}+X{row_idx}*Y{row_idx})/B{row_idx})"
                ),
            )
            ws.cell(row=row_idx, column=4, value=f"=T{row_idx}+W{row_idx}+Z{row_idx}")
            
            # E: 出清电量 ④
            ws.cell(row=row_idx, column=5, value=p.get('day_ahead', {}).get('volume', 0))
            # F: 市场均价 ⑤
            ws.cell(row=row_idx, column=6, value=p.get('day_ahead', {}).get('price', 0))
            # G: 差价电费 ⑥ = E * (F - I) [I is Price8]
            ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*(F{row_idx}-I{row_idx})")
            
            # H: 实际用量 ⑦
            ws.cell(row=row_idx, column=8, value=p.get('real_time', {}).get('volume', 0))
            # I: 市场均价 ⑧
            ws.cell(row=row_idx, column=9, value=p.get('real_time', {}).get('price', 0))
            # J: 全量电费 ⑨ = H * I
            ws.cell(row=row_idx, column=10, value=f"=H{row_idx}*I{row_idx}")
            
            # K: 电费合计 ⑩ = D + G + J
            ws.cell(row=row_idx, column=11, value=f"=D{row_idx}+G{row_idx}+J{row_idx}")
            # L: 结算均价 ⑪ = K / H
            ws.cell(row=row_idx, column=12, value=f"=IF(H{row_idx}=0, 0, K{row_idx}/H{row_idx})")
            
            # M: 机制电量 ⑫
            ws.cell(row=row_idx, column=13, value=p.get('mechanism_volume', 0))
            # N: 签约比例 ⑬ = (B + M) / H
            ws.cell(row=row_idx, column=14, value=f"=IF(H{row_idx}=0, 0, (B{row_idx}+M{row_idx})/H{row_idx})")
            # O: 标准值电费 ⑭ (Excel 分段公式)
            ws.cell(
                row=row_idx,
                column=15,
                value=(
                    f"=IF(AND(N{row_idx}>=0.8,N{row_idx}<=1.2),"
                    f"K{row_idx},"
                    f"IF(N{row_idx}<0.8,"
                    f"K{row_idx}+(0.8*H{row_idx}-(B{row_idx}+M{row_idx}))*(Q{row_idx}-I{row_idx}),"
                    f"K{row_idx}+(1.2*H{row_idx}-(B{row_idx}+M{row_idx}))*(C{row_idx}-I{row_idx})))"
                ),
            )
            # P: 结算均价 ⑮ = O / H
            ws.cell(row=row_idx, column=16, value=f"=IF(H{row_idx}=0, 0, O{row_idx}/H{row_idx})")
            # Q: 全市场合同均价 ⑯
            ws.cell(row=row_idx, column=17, value=market_prices[idx] if idx < len(market_prices) else 0)
            # R/S/T: 年度合约
            ws.cell(row=row_idx, column=18, value=year_volumes[idx] if idx < len(year_volumes) else 0)
            ws.cell(row=row_idx, column=19, value=year_prices[idx] if idx < len(year_prices) else 0)
            ws.cell(row=row_idx, column=20, value=f"=R{row_idx}*(S{row_idx}-I{row_idx})")
            # U/V/W: 月度合约
            ws.cell(row=row_idx, column=21, value=month_volumes[idx] if idx < len(month_volumes) else 0)
            ws.cell(row=row_idx, column=22, value=month_prices[idx] if idx < len(month_prices) else 0)
            ws.cell(row=row_idx, column=23, value=f"=U{row_idx}*(V{row_idx}-I{row_idx})")
            # X/Y/Z: 月内合约
            ws.cell(row=row_idx, column=24, value=intra_volumes[idx] if idx < len(intra_volumes) else 0)
            ws.cell(row=row_idx, column=25, value=intra_prices[idx] if idx < len(intra_prices) else 0)
            ws.cell(row=row_idx, column=26, value=f"=X{row_idx}*(Y{row_idx}-I{row_idx})")

            # 设置数字格式和边框
            for col in range(1, 27):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                if col > 1:
                    if col == 14:
                        cell.number_format = '0.0%'
                    elif col in {2, 5, 8, 13, 18, 21, 24}:
                        cell.number_format = '0.000'
                    else:
                        cell.number_format = '0.00'

        # 5. 合计行
        total_row_idx = 4 + len(details)
        ws.cell(row=total_row_idx, column=1, value="合计").font = header_font
        for col in [2, 4, 5, 7, 8, 10, 11, 13, 15, 18, 20, 21, 23, 24, 26]:
            col_letter = get_column_letter(col)
            ws.cell(row=total_row_idx, column=col, value=f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})")

        # 合计行的均价/比例/参考价格
        ws.cell(
            row=total_row_idx,
            column=3,
            value=(
                f"=IF(B{total_row_idx}=0,0,"
                f"(R{total_row_idx}*S{total_row_idx}+U{total_row_idx}*V{total_row_idx}+X{total_row_idx}*Y{total_row_idx})/B{total_row_idx})"
            ),
        )
        ws.cell(row=total_row_idx, column=12, value=f"=IF(H{total_row_idx}=0, 0, K{total_row_idx}/H{total_row_idx})")
        ws.cell(row=total_row_idx, column=14, value=f"=IF(H{total_row_idx}=0, 0, (B{total_row_idx}+M{total_row_idx})/H{total_row_idx})")
        ws.cell(row=total_row_idx, column=16, value=f"=IF(H{total_row_idx}=0, 0, O{total_row_idx}/H{total_row_idx})")
        ws.cell(row=total_row_idx, column=17, value=market_contract_avg.get("daily_avg_price", 0))
        ws.cell(row=total_row_idx, column=19, value=yearly_contract.get("daily_avg_price", 0))
        ws.cell(row=total_row_idx, column=22, value=monthly_contract.get("daily_avg_price", 0))
        ws.cell(row=total_row_idx, column=25, value=intra_month_contract.get("daily_avg_price", 0))

        # 合计行样式
        for col in range(1, 27):
            cell = ws.cell(row=total_row_idx, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
            cell.border = thin_border
            if col > 1:
                if col == 14:
                    cell.number_format = '0.0%'
                elif col in {2, 5, 8, 13, 18, 21, 24}:
                    cell.number_format = '0.000'
                else:
                    cell.number_format = '0.00'

        # 列宽微调
        for col in range(1, 27):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
